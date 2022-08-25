import openpyxl

# wb=openpyxl.Workbook()

wb = openpyxl.load_workbook("Book1.xlsx")

print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_sheet("Sheet2",2)
# wb.remove_sheet()

# cell = sheet["a1"]
# print(cell.value)
# cell.value = 1
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

# cell_two = sheet.cell(row=1, column=1)
# print(cell_two.value)
# print(sheet.max_row)
# print(sheet.max_column)


# for row in range(1, sheet.max_row+1):
#     for column in range(1, sheet.max_column+1):
#         cell_three = sheet.cell(row, column)
#         print(cell_three.value)

sheet.append([1, 2, 3])

wb.save("transaction.xlsx")
